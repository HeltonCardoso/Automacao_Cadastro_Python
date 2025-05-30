import os
import warnings
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tkinter.ttk import Progressbar
from datetime import datetime
from openpyxl.styles import NamedStyle

# Suprimir o warning específico do openpyxl
warnings.filterwarnings("ignore", message="Data Validation extension is not supported and will be removed")

# Funções para copiar e reaplicar as validações de dados
from openpyxl.worksheet.datavalidation import DataValidation

def copiar_validacoes(worksheet):
    """ Copia todas as validações de dados da planilha """
    return list(worksheet.data_validations.dataValidation)

def reaplicar_validacoes(worksheet, validacoes):
    """ Reaplica as validações de dados na planilha """
    for dv in validacoes:
        worksheet.add_data_validation(dv)

class TelaCadastroProduto:
    def __init__(self, root):
        self.root = root
        self.root.title("CADASTRO PLANILHA ATHUS V1.0")
        self.root.resizable(False, False)
        self.centralizar_janela(750, 520)
        self.aplicar_estilo_global()  # Aplica o estilo global

        # Adicionar ícone à janela
        try:
            self.root.iconbitmap("IMG/icone.ico")  # Substitua "icone.ico" pelo caminho do seu ícone
        except:
            pass  # Ignora se o ícone não for encontrado

        # Frame principal
        frame = tk.Frame(self.root)
        frame.pack(pady=20)

        # Label e Entry para selecionar a planilha online
        tk.Label(frame, text="PLANILHA ONLINE:").grid(row=0, column=0)
        self.entrada_origem = tk.Entry(frame, width=70, font=("Arial", 10))
        self.entrada_origem.grid(row=0, column=1)
        btn_buscar_origem = tk.Button(
            frame, 
            text="SELECIONE", 
            command=lambda: self.selecionar_arquivo(self.entrada_origem), 
            bg="#008CBA", 
            fg="white", 
            font=("Arial", 10, "bold"), 
            bd=4, 
            relief=tk.FLAT
        )
        btn_buscar_origem.grid(row=0, column=2)
        btn_buscar_origem.bind("<Enter>", lambda e: btn_buscar_origem.config(bg="#0a2040"))  # Efeito hover
        btn_buscar_origem.bind("<Leave>", lambda e: btn_buscar_origem.config(bg="#008CBA"))  # Efeito ao sair

        # Label e Entry para selecionar o modelo Athus
        tk.Label(frame, text="PLANILHA ATHUS:").grid(row=1, column=0)
        self.entrada_destino = tk.Entry(frame, width=70, font=("Arial", 10))
        self.entrada_destino.grid(row=1, column=1)
        
        btn_buscar_destino = tk.Button(
            frame, 
            text="SELECIONE", 
            command=lambda: self.selecionar_arquivo(self.entrada_destino), 
            bg="#008CBA", 
            fg="white", 
            font=("Arial", 10, "bold"), 
            bd=4, 
            relief=tk.FLAT
        )
        btn_buscar_destino.grid(row=1, column=2, pady=(10, 10))
        btn_buscar_destino.bind("<Enter>", lambda e: btn_buscar_destino.config(bg="#0a2040"))  # Efeito hover
        btn_buscar_destino.bind("<Leave>", lambda e: btn_buscar_destino.config(bg="#008CBA"))  # Efeito ao sair

        # Botão "IMPORTAR PLANILHA ONLINE"
        btn_importar = tk.Button(
            self.root, 
            text="IMPORTAR PLANILHA ONLINE", 
            command=self.processar, 
            bg="#008CBA", 
            fg="white", 
            font=("Arial", 10, "bold"), 
            bd=4, 
            relief=tk.FLAT
        )
        btn_importar.pack(pady=10)
        btn_importar.bind("<Enter>", lambda e: btn_importar.config(bg="#0a2040"))  # Efeito hover
        btn_importar.bind("<Leave>", lambda e: btn_importar.config(bg="#008CBA"))  # Efeito ao sair

        # Barra de progresso
        self.progresso = Progressbar(self.root, orient="horizontal", length=550, mode="determinate")
        self.progresso.pack(pady=5)

        # Label de status
        self.status_label = tk.Label(self.root, text="Pronto para iniciar...", fg="gray", font=("Arial", 10))
        self.status_label.pack(pady=5)

        # Área de logs
        self.log_area = scrolledtext.ScrolledText(
            self.root, 
            width=80, 
            height=10, 
            bg="white", 
            state="disabled", 
            font=("Courier", 10)
        )
        self.log_area.pack(pady=10)

        # Configuração de tags para cores
        self.log_area.tag_config("info", foreground="blue", font=("Courier", 11, "bold"))
        self.log_area.tag_config("erro", foreground="red", font=("Courier", 11, "bold"))
        self.log_area.tag_config("sucesso", foreground="green", font=("Courier", 11, "bold"))
        self.log_area.tag_config("aviso", foreground="orange", font=("Courier", 11, "bold"))

        # Botão "FECHAR"
        btn_fechar = tk.Button(
            self.root, 
            text="FECHAR", 
            command=self.root.destroy, 
            bg="#f44336", 
            fg="white", 
            font=("Arial", 10, "bold"), 
            bd=4, 
            relief=tk.FLAT
        )
        btn_fechar.pack(pady=10)
        btn_fechar.bind("<Enter>", lambda e: btn_fechar.config(bg="#e53935"))  # Efeito hover
        btn_fechar.bind("<Leave>", lambda e: btn_fechar.config(bg="#f44336"))  # Efeito ao sair

        # Rodapé
        frame_rodape = tk.Frame(root)
        frame_rodape.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

        # Texto de versão
        texto_versao = tk.Label(frame_rodape, text="Versão 1.0 - Desenvolvido por Helton", fg="gray", font=("Arial", 8))
        texto_versao.pack(side=tk.LEFT, padx=10)

        # Relógio no rodapé
        self.relogio = tk.Label(frame_rodape, font=("Arial", 8), fg="gray")
        self.relogio.pack(side=tk.RIGHT, padx=10)
        self.atualizar_relogio()

    def aplicar_estilo_global(self):
        """Aplica o estilo global da interface."""
        style = ttk.Style(self.root)
        style.configure("TButton", font=("Arial", 10, "bold"), padding=10, relief="flat", background="#008CBA", foreground="white")
        style.map("TButton", background=[("active", "#0a2040")], foreground=[("active", "white")])
        style.configure("Fechar.TButton", background="#f44336", foreground="white")
        style.map("Fechar.TButton", background=[("active", "#e53935")], foreground=[("active", "white")])
        style.configure("TLabel", font=("Arial", 10), foreground="gray")
        style.configure("TEntry", font=("Arial", 10), relief="flat")

    def centralizar_janela(self, largura, altura):
        """Centraliza a janela na tela."""
        largura_tela = self.root.winfo_screenwidth()
        altura_tela = self.root.winfo_screenheight()
        pos_x = (largura_tela // 2) - (largura // 2)
        pos_y = (altura_tela // 2) - (altura // 2)
        self.root.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")

    def atualizar_relogio(self):
        """Atualiza o relógio no rodapé."""
        agora = datetime.now()
        data_hora = agora.strftime("%d/%m/%Y %H:%M:%S")
        self.relogio.config(text=data_hora)
        self.root.after(1000, self.atualizar_relogio)  # Atualiza a cada 1 segundo

    def selecionar_arquivo(self, entrada):
        """Abre o diálogo para selecionar um arquivo."""
        self.root.attributes('-topmost', False)  # Desativa o topmost temporariamente
        caminho = filedialog.askopenfilename(
            parent=self.root,  # Define a janela pai como a janela de cadastro
            filetypes=(("Arquivos Excel", "*.xlsx *.xls *.csv"), ("Todos os arquivos", "*.*"))
        )
        self.root.attributes('-topmost', True)  # Reativa o topmost

        if caminho:
            entrada.delete(0, tk.END)
            entrada.insert(0, caminho)

    def log(self, mensagem, tipo="info"):
        """Adiciona uma mensagem à área de logs com timestamp."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        mensagem_formatada = f"[{timestamp}] {mensagem}\n"

        self.log_area.config(state="normal")
        self.log_area.insert(tk.END, mensagem_formatada, tipo)
        self.log_area.config(state="disabled")
        self.log_area.yview(tk.END)  # Rola para a última linha

    def processar(self):
        """Processa as planilhas selecionadas."""
        planilha_origem = self.entrada_origem.get()
        planilha_destino = self.entrada_destino.get()

        if not planilha_origem or not planilha_destino:
            messagebox.showwarning("Atenção", "Selecione os arquivos de origem e destino!", parent=self.root)
            return

        try:
            caminho_arquivo_salvo = self.executar_processamento(planilha_origem, planilha_destino)
            
            # Criar uma janela de mensagem personalizada
            msg_box = tk.Toplevel(self.root)
            msg_box.title("Sucesso")
            msg_box.resizable(False, False)
            msg_box.transient(self.root)  # Define como janela filha
            msg_box.grab_set()  # Modal
            
            # Centralizar a janela de mensagem
            largura_janela = 500
            altura_janela = 150
            pos_x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (largura_janela // 2)
            pos_y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (altura_janela // 2)
            msg_box.geometry(f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}")
            
            # Adicionar conteúdo
            tk.Label(msg_box, text="Cadastro realizado com sucesso!", font=("Arial", 12)).pack(pady=10)
            tk.Label(msg_box, text=f"Arquivo salvo em:\n{caminho_arquivo_salvo}", wraplength=450).pack()
            
            # Frame para os botões
            frame_botoes = tk.Frame(msg_box)
            frame_botoes.pack(pady=10)
            
            # Botão OK
            btn_ok = tk.Button(
                frame_botoes, 
                text="OK", 
                command=msg_box.destroy,
                width=10
            )
            btn_ok.pack(side=tk.LEFT, padx=5)
            
            # Botão Abrir Arquivo
            btn_abrir = tk.Button(
                frame_botoes, 
                text="Abrir Arquivo", 
                command=lambda: self.abrir_arquivo(caminho_arquivo_salvo, msg_box),
                width=10
            )
            btn_abrir.pack(side=tk.LEFT, padx=5)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro durante o processamento: {e}", parent=self.root)

    def mostrar_janela_salvar(self, titulo, tipos_arquivo, nome_inicial):
        """Exibe a janela de salvamento garantindo que fique na frente"""
        # Esconde temporariamente a janela principal
        self.root.withdraw()
        
        # Abre a janela de salvamento
        caminho = filedialog.asksaveasfilename(
            title=titulo,
            defaultextension=".xlsx",
            filetypes=tipos_arquivo,
            initialfile=nome_inicial
        )
        
        # Restaura a janela principal
        self.root.deiconify()
        self.root.lift()
        self.root.focus_force()
        
        return caminho
    
    def abrir_arquivo(self, caminho_arquivo, janela_pai=None):
                """Abre o arquivo no aplicativo padrão do sistema."""
                try:
                    if os.path.exists(caminho_arquivo):
                        os.startfile(caminho_arquivo)  # Para Windows
                    else:
                        messagebox.showwarning("Aviso", "O arquivo não foi encontrado!", parent=self.root)
                    
                    if janela_pai:
                        janela_pai.destroy()
                except Exception as e:
                    messagebox.showerror("Erro", f"Não foi possível abrir o arquivo: {e}", parent=self.root)
                    if janela_pai:
                        janela_pai.destroy()
    

    def executar_processamento(self, planilha_origem, planilha_destino):
        """Executa o processamento das planilhas."""
        try:
            # Ler os dados da planilha de origem
            df = pd.read_excel(planilha_origem)
            total_linhas = len(df)
            self.progresso["maximum"] = total_linhas
            self.progresso["value"] = 0

            # Criar listas de DataFrames para cada aba
            dados_sheets = {
                "PRODUTO": [],
                "PRECO": [],
                "LOJA WEB": [],
                "KIT": [],
                "VOLUME": []
            }

            # Criar dicionário {EAN: Nome do Produto}
            produto_dict = {
                str(row["EAN"]).strip(): row["NOMEONCLICK"] if pd.notna(row["NOMEONCLICK"]) else "Nome Desconhecido"
                for _, row in df.iterrows()
            }

            # Criar conjunto para armazenar a marca cadastrada
            marcas_cadastradas = set()
            data_atual = datetime.now()

            # Formatar a data atual no formato desejado (DD/MM/YYYY)
            data_formatada = data_atual.strftime("%d/%m/%Y")

            # Adicionar 30 anos à data atual
            data_mais_20_anos = data_atual.replace(year=data_atual.year + 30)

            # Formatar a nova data no mesmo formato
            data_formatada_mais_20_anos = data_mais_20_anos.strftime("%d/%m/%Y")
        
            for idx, row in df.iterrows():
                ean = str(row["EAN"]).strip()
                tipo_produto = row["TIPODEPRODUTO"].strip().upper()
                #descricao = row["PRODUTO"]
                #quantidade = row["QTDECOMPONENTES"]
                #cor = row["COR"]
                altura = row["EMBALTURA"]
                largura = row["EMBLARGURA"]
                comprimento = row["EMBCOMPRIMENTO"]
                volumes = int(row["VOLUMES"]) if pd.notna(row["VOLUMES"]) else 1  # Garante que tenha pelo menos 1 volume
                componentes = row["EANCOMPONENTES"]
                marca = row["MARCA"]
                custo = row["CUSTO"]
                preco_venda = row["DE"]
                preco_promo = row["POR"]
                fornecedor = row["FORNECEDOR"]
                outros = row["OUTROS"]
                frete  = row["FRETE"]
                ncm = row["NCM"]
                cod_forn = row["CODFORN"]
                nome_onclick = row["NOMEONCLICK"]
                categoria = row["CATEGORIA"]
                grupo = row["GRUPO"]
                nome_ecommerce = row["NOMEE-COMMERCE"]
                marca_web = ""
                if isinstance(nome_ecommerce, str) and "-" in nome_ecommerce:
                # Pega a última parte após o último hífen e remove espaços
                    marca_web = nome_ecommerce.split("-")[-1].strip()
                    
                complemento = row["COMPLEMENTO"]
                disponibilidade_web = row["DISPONIBILIDADEWEB"]
                descricao_html = row["DESCRICAOHTML"]
                peso_bruto = row["PESOBRUTO"]
                peso_liquido = row["PESOLIQUIDO"]
                vol_peso_bruto = row["VOLPESOBRUTO"]
                vol_peso_liquido = row["VOLPESOLIQ"]
                vol_largura = row["VOLLARGURA"]
                vol_altura = row["VOLALTURA"]
                vol_comprimento = row["VOLCOMPRIMENTO"]
                tipo_produto_valor = 0 if tipo_produto == "PRODUTO ACABADO" else 2
                # Regra: Nome do produto (5ª coluna) com apenas 25 caracteres
                nome_reduzido = nome_onclick[:25] if isinstance(nome_onclick, str) else ""
                # Adicionar marca ao conjunto
                marcas_cadastradas.add(marca)
                # Usar sempre as medidas PESOBRUTO, PESOLIQUIDO, EMBLARGURA, EMBALTURA, EMBCOMPRIMENTO para a aba "PRODUTO"
                peso_bruto_final = peso_bruto
                peso_liquido_final = peso_liquido
                largura_final = largura
                altura_final = altura
                comprimento_final = comprimento

                # Adicionar dados ao dicionário de listas
                dados_sheets["PRODUTO"].append([
                    ean, cod_forn, tipo_produto_valor, nome_onclick, nome_reduzido, nome_onclick, nome_onclick, None,
                    marca, categoria, grupo, None, None, complemento, None, None, "F", "F", "F", None, volumes,
                    peso_bruto_final, peso_liquido_final, largura_final, altura_final, comprimento_final, None, 90, 1000,
                    disponibilidade_web, "F", "F", ncm, None, "0", "T", "F", "F", "NAO", nome_ecommerce, marca_web,
                    "90 dias após o recebimento do produto", disponibilidade_web, descricao_html, "F", "F"
                ])

                if tipo_produto == "KIT" and pd.notna(componentes):
                    for comp in str(componentes).split("/"):
                        comp_ean = comp.strip()
                        nome_componente = produto_dict.get(comp_ean, "Desconhecido")
                        dados_sheets["KIT"].append([ean, comp_ean, nome_componente, "1", "", "0"])

                # Adicionar volumes (Agora inclui produtos com apenas 1 volume)
                for i in range(volumes):
                    if volumes == 1:
                        # Quando for 1 volume, usar dados da aba "PRODUTO"
                        dados_sheets["VOLUME"].append([
                            ean, nome_onclick, peso_bruto_final, peso_liquido_final, largura_final, altura_final, "",
                            comprimento_final, "", "BOX", "T", i + 1
                        ])
                    else:
                        # Quando houver mais de 1 volume, usar os dados de volume específico
                        dados_sheets["VOLUME"].append([
                            ean, nome_onclick, vol_peso_bruto, vol_peso_liquido, vol_largura, vol_altura, "",
                            vol_comprimento, "", "BOX", "T", i + 1
                        ])

                dados_sheets["PRECO"].append([
                    ean, fornecedor, custo, outros, "", frete, row["CUSTOTOTAL"], preco_venda, preco_promo, preco_promo,
                    data_formatada, data_formatada_mais_20_anos, "", "F"
                ])

                dados_sheets["LOJA WEB"].append([
                    ean, "", "", "", row["CATEGORIAPRINCIPALTRAY"], "", "", "", "T", "F", "", "", "",
                    row["CATEGORIAPRINCIPALCORP"], row["NIVELADICIONAL1CORP"], "", "", "T", "T"
                ])

                # Atualizar a barra de progresso e o status
                self.progresso["value"] = idx + 1
                self.status_label.config(text=f"Processando linha {idx + 1} de {total_linhas}...")
                self.log(f"Processando produto: {nome_onclick}")
                self.root.update_idletasks()

            # Carregar a planilha original mantendo formatação
            wb = load_workbook(planilha_destino)
            # Cria um estilo personalizado "aspas_invisiveis"
            estilo_invisivel = NamedStyle(name="aspas_invisiveis")
            estilo_invisivel.number_format = '@'  # Formato texto
            wb.add_named_style(estilo_invisivel)

            # Obter a aba "Tipo Importacao" e copiar as validações antes de salvar
            ws_tipo_importacao = wb["Tipo Importacao"]
            validacoes_tipo_importacao = copiar_validacoes(ws_tipo_importacao)

            # Sobrescrever os dados mantendo formatação
            for sheet_name, data in dados_sheets.items():
                ws = wb[sheet_name]

                # Definir linha inicial dependendo da aba
                if sheet_name == "PRODUTO":
                    start_row = 3  # Para a aba PRODUTO, começa na linha 3
                else:
                    start_row = 2  # Para as outras abas, começa na linha 2

                # Limpar apenas os dados (não os cabeçalhos)
                for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None

                # Escrever os novos dados
                for i, row_data in enumerate(data, start=start_row):
                    for j, value in enumerate(row_data, start=1):
                        ws.cell(row=i, column=j, value=value)
            
            # [6] --- BLOCO DE PÓS-PROCESSAMENTO (NOVO) ---
            for row in ws.iter_rows(min_row=start_row):
                for cell in row:
                    if cell.value == "'":
                        # Aplica o estilo especial
                        cell.style = "aspas_invisiveis"
                        # Força o Excel a tratar como texto literal
                        cell.value = "'"
                        # Remove qualquer formatação visual
                        cell.fill = None    # Sem preenchimento


            # Gerar o nome do novo arquivo com a marca
            if marcas_cadastradas:
                marca_unica = next(iter(marcas_cadastradas))  # Obtém a única marca do conjunto
                novo_nome_arquivo = f"Template_Produtos_Mpozenato_CADASTRO_{marca_unica}.xlsx"

                # Abrir caixa de diálogo para selecionar o local de salvamento
                # Substitua a parte do salvamento por:
                caminho_arquivo = self.mostrar_janela_salvar(
                    titulo="Salvar Planilha Processada",
                    tipos_arquivo=[("Arquivos Excel", "*.xlsx")],
                    nome_inicial=novo_nome_arquivo
                )

                if not caminho_arquivo:
                    self.log("Operação de salvamento cancelada pelo usuário.", "aviso")
                    return

                try:
                    wb.save(caminho_arquivo)
                    self.log("Arquivo salvo com sucesso!")
                except PermissionError:
                    messagebox.showerror("Erro", "Feche o arquivo de saída e tente novamente!", parent=self.root)
                    return

                # Reabrir a planilha e reaplicar as validações na aba "Tipo Importação"
                wb = load_workbook(caminho_arquivo)
                ws_tipo_importacao = wb["Tipo Importacao"]
                reaplicar_validacoes(ws_tipo_importacao, validacoes_tipo_importacao)

                # Salvar novamente após reaplicar as validações
                wb.save(caminho_arquivo)

            return caminho_arquivo  # Retorna o caminho completo do arquivo salvo

        except Exception as e:
            self.log(f"Erro durante o processamento: {e}", "erro")
            raise e

# Função principal
if __name__ == "__main__":
    root = tk.Tk()
    app = TelaCadastroProduto(root)
    root.mainloop()